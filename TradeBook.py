import sys
import sqlite3
import openpyxl
import os


class TradeBook():
    def __init__(self, db_path, db_filename,trade_table):
        self.trades = []
        self.cursor = None
        self.db_path = db_path
        self.db_filename = db_filename
        self.trade_table = trade_table
        return

    def readtrades(self, path, filename):
        # reads trades from wbk and stores it in local memory (as a list of tuples)
        # keep only those in memory that are not there in database Trades table
        try:
            wb = openpyxl.load_workbook(os.path.join(path, filename))
        except FileNotFoundError:
            sys.stderr.write('FATAL ERROR:\nTrade files could not be loaded! Either the file doesn\'t exist or is corrupted\n')
            return False

        ws = wb['TradeBook']
        try:
            connexion = sqlite3.connect(os.path.join(self.db_path, self.db_filename))
            self.cursor = connexion.cursor()
            row = self.cursor.execute("SELECT name FROM sqlite_master where type =:type_name and name =:table ",
                                    {'type_name': 'table', 'table': self.trade_table}).fetchall()
            # first pass of trade, no subsequent trade filtering to be done
            if len(row) == 0:
                for row in ws.iter_rows(min_row=2, max_col=10, max_row=100, values_only=True):
                    if row[6] is not None:
                        self.trades.append(tuple([row[0], row[1], row[2], row[3], row[4], row[6], row[9]]))

            else:
                res = self.cursor.execute('''select TradeId from ''' + self.trade_table + ''' where TradeId is not NULL''').fetchall()
                res = [i[0] for i in res]
                for row in ws.iter_rows(min_row=2, max_col=10, max_row=100, values_only=True):
                    if (row[6] not in res) and (row[6] is not None):
                        self.trades.append(tuple([row[0], row[1], row[2], row[3], row[4], row[6], row[9]]))


        except sqlite3.OperationalError:
            pass

        return True

    def write_db(self):
        # takes a list of tuples, writes into db
        try:
            connexion = sqlite3.connect(os.path.join(self.db_path, self.db_filename))
            self.cursor = connexion.cursor()
            if len(self.trades) == 0:
                sys.stderr.write('CAUTION!:\nNo prior trades loaded')
                raise ValueError

            trade_fields = [('TradeDate', 'text'), ('Ticker', 'text'), ('Action', 'text'), ('Qty', 'integer'),
                            ('Price', 'real'),('TradeId','text'),('Demat','text')]
            self.__table_exists('TradeBook', trade_fields, connexion)

            for i in range(len(self.trades)):
                self.cursor.execute(
                    '''INSERT INTO TradeBook ('TradeDate','Ticker','Action','Qty','Price','TradeId','Demat') 
                    VALUES (?,?,?,?,?,?,?)''',
                    self.trades[i])
            connexion.commit()

        except sqlite3.OperationalError:
            sys.stderr.write('CAUTION!:\nEither the database doesn\'t exist or the path is improper')
            return
        except ValueError:
            sys.stderr.write('No prior trades loaded and file path and name offered is invalid')
            return

    def __table_exists(self, tablename, fields, connector):

        row = self.cursor.execute("SELECT name FROM sqlite_master where type =:type_name and name =:table ",
                                  {'type_name': 'table', 'table': tablename}).fetchall()
        if len(row) == 0:
            sys.stderr.write('No such table exists in the database.\n')
            sys.stderr.write('Creating a new table in the database.\n')
            fields = [i[0] + ' ' + i[1] for i in fields]
            fields = ','.join(fields)

            self.cursor.execute(
                '''CREATE  TABLE ''' + tablename + ''' (''' + fields + ''')''')
            connector.commit()

        return

    def min(self, field: str):
        try:
            idx = ['Date','Ticker','Value','Qty','Price'].index(field)
        except ValueError:
            sys.stderr.write('Field not found. Choose between Date,Ticker,Value,Qty,Price')
        if len(self.trades)!=0:
            for i in self.trades:
                if idx !=2:
                    _ = [i[idx] for i in self.trades]
                else:
                    _ = [i[3]*i[4] for i in self.trades]

            _.sort(reverse = False)
            return _[0]
        else:
            return None

